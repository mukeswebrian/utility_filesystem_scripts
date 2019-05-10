import PyPDF2
from openpyxl import load_workbook

wb = load_workbook('reference List.xlsx')
ws = wb['Other references']

def extract_files(start, end, sourceFile=''):
    pdfFile = open(sourceFile, 'rb')
    end = end+1

    for i in range(start,end):
        pdfReader = PyPDF2.PdfFileReader(pdfFile)
        pdfWriter = PyPDF2.PdfFileWriter()
        strt_page = ws['C'+str(i)].value
        strt_page = strt_page - 1

        for pageNum in range(strt_page,ws['D'+str(i)].value):
            pageObj = pdfReader.getPage(pageNum)
            pdfWriter.addPage(pageObj)

        outname = str(ws['A'+str(i)].value)
        outname = outname.replace('PUB/','PUB-')

        pdfOutputFile = open(outname+'.pdf', 'wb')
        pdfWriter.write(pdfOutputFile)
        pdfOutputFile.close()
        print('Successfully extracted file: '+outname)

    pdfFile.close()
    print('Done extracting from: '+sourceFile)

#extract_files(3, 17, sourceFile='round_1_pub_irs.pdf')
#extract_files(18, 46, sourceFile='round_2_pub_irs.pdf')
#extract_files(2, 15, sourceFile='full_rate_application_2017.pdf')
#extract_files(16, 35, sourceFile='pub_mfrs revised June 21.pdf')
#extract_files(36, 58, sourceFile='round_1_pub_irs.pdf')


'''
pdfFile = open('round_1_pub_irs.pdf', 'rb')

for i in range(3,17):
    pdfReader = PyPDF2.PdfFileReader(pdfFile)
    pdfWriter = PyPDF2.PdfFileWriter()

    for pageNum in range(ws['C'+str(i)].value,ws['D'+str(i)].value):
        pageObj = pdfReader.getPage(pageNum)
        pdfWriter.addPage(pageObj)

    outname = str(ws['A'+str(i)].value)
    outname = outname.replace('PUB/','PUB-')

    pdfOutputFile = open(outname+'.pdf', 'wb')
    pdfWriter.write(pdfOutputFile)
    pdfOutputFile.close()
    print('Successfully extracted file: '+outname)

pdfFile.close()
print('Done IR round one \n')

pdfFile = open('round_2_pub_irs.pdf', 'rb')

for i in range(18,46):
    pdfReader = PyPDF2.PdfFileReader(pdfFile)
    pdfWriter = PyPDF2.PdfFileWriter()

    for pageNum in range(ws['C'+str(i)].value,ws['D'+str(i)].value):
        pageObj = pdfReader.getPage(pageNum)
        pdfWriter.addPage(pageObj)

    outname = str(ws['A'+str(i)].value)
    outname = outname.replace('PUB/','PUB-')

    pdfOutputFile = open(outname+'.pdf', 'wb')
    pdfWriter.write(pdfOutputFile)
    pdfOutputFile.close()
    print('Successfully extracted file: '+outname)

pdfFile.close()
print('Done IR round 2')
'''
